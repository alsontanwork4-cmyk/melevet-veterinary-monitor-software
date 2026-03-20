from __future__ import annotations

from datetime import UTC, datetime, timedelta
from io import BytesIO
import time
import zipfile

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session, sessionmaker

from app.database import Base
from app.models import (
    Channel,
    Measurement,
    NibpEvent,
    Patient,
    RecordingPeriod,
    Segment,
    SourceType,
    Upload,
    UploadStatus,
)
from app.routers.decode import router as decode_router
from app.services import decode_job_service
from app.services.decode_export_service import (
    NIBP_WORKBOOK_NAME,
    TREND_WORKBOOK_NAME,
    build_decode_export_archive,
)


app = FastAPI()
app.include_router(decode_router, prefix="/api")
client = TestClient(app)


@pytest.fixture(autouse=True)
def clear_decode_jobs() -> None:
    with decode_job_service._lock:
        decode_job_service._jobs.clear()
    yield
    with decode_job_service._lock:
        decode_job_service._jobs.clear()


def _file_bytes(data_dir, relative_path: str) -> bytes:
    return (data_dir / relative_path).read_bytes()


def _zip_names(payload: bytes) -> list[str]:
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        return sorted(archive.namelist())


def _sheet_headers_from_zip(payload: bytes, workbook_name: str, sheet_name: str) -> list[str]:
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        with archive.open(workbook_name) as workbook_file:
            workbook = load_workbook(BytesIO(workbook_file.read()), read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    return [str(cell) if cell is not None else "" for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]


def _wait_for_job(job_id: str, *, timeout_seconds: float = 30.0) -> dict[str, object]:
    started_at = time.monotonic()
    while time.monotonic() - started_at < timeout_seconds:
        response = client.get(f"/api/decode/jobs/{job_id}")
        assert response.status_code == 200
        payload = response.json()
        if payload["status"] in {"completed", "error"}:
            return payload
        time.sleep(0.05)
    raise AssertionError(f"Timed out waiting for decode job {job_id}")


def _build_session() -> Session:
    engine = create_engine("sqlite+pysqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, class_=Session)
    return session_factory()


def _table_counts(db: Session) -> dict[str, int]:
    return {
        "patients": db.scalar(select(func.count(Patient.id))) or 0,
        "uploads": db.scalar(select(func.count(Upload.id))) or 0,
        "recording_periods": db.scalar(select(func.count(RecordingPeriod.id))) or 0,
        "segments": db.scalar(select(func.count(Segment.id))) or 0,
        "channels": db.scalar(select(func.count(Channel.id))) or 0,
        "measurements": db.scalar(select(func.count(Measurement.id))) or 0,
        "nibp_events": db.scalar(select(func.count(NibpEvent.id))) or 0,
    }


def _seed_decode_irrelevant_rows(db: Session) -> None:
    checksum = "1" * 64
    patient = Patient(patient_id_code="DECODE-001", name="Decode Patient", species="Dog")
    db.add(patient)
    db.flush()

    upload = Upload(
        patient_id=patient.id,
        status=UploadStatus.completed,
        phase="completed",
        progress_current=1,
        progress_total=1,
        trend_frames=1,
        nibp_frames=1,
        trend_sha256=checksum,
        trend_index_sha256=checksum,
        nibp_sha256=checksum,
        nibp_index_sha256=checksum,
        combined_hash=checksum,
        detected_local_dates=["2026-03-07"],
        completed_at=datetime(2026, 3, 7, 0, 0, 0, tzinfo=UTC),
    )
    db.add(upload)
    db.flush()

    period = RecordingPeriod(
        upload_id=upload.id,
        period_index=0,
        start_time=datetime(2026, 3, 7, 0, 0, 0),
        end_time=datetime(2026, 3, 7, 0, 5, 0),
        frame_count=1,
        label="Period 1",
    )
    db.add(period)
    db.flush()

    segment = Segment(
        period_id=period.id,
        upload_id=upload.id,
        segment_index=0,
        start_time=period.start_time,
        end_time=period.end_time,
        frame_count=1,
        duration_seconds=300,
    )
    db.add(segment)
    db.flush()

    channel = Channel(
        upload_id=upload.id,
        source_type=SourceType.trend,
        channel_index=16,
        name="heart_rate_be_u16",
        unit="bpm",
        valid_count=1,
    )
    db.add(channel)
    db.flush()

    db.add(
        Measurement(
            upload_id=upload.id,
            segment_id=segment.id,
            channel_id=channel.id,
            timestamp=datetime(2026, 3, 7, 0, 1, 0),
            value=88.0,
            dedup_key="measurement-seed-1",
        )
    )
    db.add(
        NibpEvent(
            upload_id=upload.id,
            segment_id=segment.id,
            timestamp=datetime(2026, 3, 7, 0, 2, 0),
            channel_values={"bp_systolic_inferred": 120, "bp_mean_inferred": 90, "bp_diastolic_inferred": 70},
            has_measurement=True,
            dedup_key="nibp-seed-1",
        )
    )
    db.commit()


def test_decode_export_rejects_empty_request() -> None:
    response = client.post("/api/decode/export", data={"timezone": "UTC"})

    assert response.status_code == 400
    assert response.json()["detail"] == "Upload at least one complete record pair to decode"


def test_decode_export_rejects_partial_pair(data_dir) -> None:
    response = client.post(
        "/api/decode/export",
        data={"timezone": "UTC"},
        files={
            "trend_data": ("TrendChartRecord.data", _file_bytes(data_dir, "Records/TrendChartRecord.data"), "application/octet-stream"),
        },
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Trend files must be uploaded together: trend_data + trend_index"


def test_decode_export_returns_expected_zip_for_single_pair(data_dir) -> None:
    response = client.post(
        "/api/decode/export",
        data={"timezone": "Asia/Singapore"},
        files={
            "trend_data": ("TrendChartRecord.data", _file_bytes(data_dir, "Records/TrendChartRecord.data"), "application/octet-stream"),
            "trend_index": ("TrendChartRecord.Index", _file_bytes(data_dir, "Records/TrendChartRecord.Index"), "application/octet-stream"),
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/zip"
    assert "decoded-records.zip" in response.headers["content-disposition"]
    assert _zip_names(response.content) == [TREND_WORKBOOK_NAME]


def test_decode_export_returns_expected_zip_for_all_supported_pairs(data_dir) -> None:
    response = client.post(
        "/api/decode/export",
        data={"timezone": "Asia/Singapore"},
        files={
            "trend_data": ("TrendChartRecord.data", _file_bytes(data_dir, "Records/TrendChartRecord.data"), "application/octet-stream"),
            "trend_index": ("TrendChartRecord.Index", _file_bytes(data_dir, "Records/TrendChartRecord.Index"), "application/octet-stream"),
            "nibp_data": ("NibpRecord.data", _file_bytes(data_dir, "Records/NibpRecord.data"), "application/octet-stream"),
            "nibp_index": ("NibpRecord.Index", _file_bytes(data_dir, "Records/NibpRecord.Index"), "application/octet-stream"),
        },
    )

    assert response.status_code == 200
    assert _zip_names(response.content) == sorted([TREND_WORKBOOK_NAME, NIBP_WORKBOOK_NAME])


def test_decode_job_reports_progress_and_supports_download(data_dir) -> None:
    response = client.post(
        "/api/decode/jobs",
        data={"timezone": "Asia/Singapore"},
        files={
            "trend_data": ("TrendChartRecord.data", _file_bytes(data_dir, "Records/TrendChartRecord.data"), "application/octet-stream"),
            "trend_index": ("TrendChartRecord.Index", _file_bytes(data_dir, "Records/TrendChartRecord.Index"), "application/octet-stream"),
        },
    )

    assert response.status_code == 202
    payload = response.json()
    assert payload["id"]
    assert payload["selected_families"] == ["Trend Chart"]
    assert payload["status"] in {"queued", "processing", "completed"}

    final_payload = _wait_for_job(payload["id"])
    assert final_payload["status"] == "completed"
    assert final_payload["progress_percent"] == 100
    assert final_payload["phase"] == "Ready to download"

    download_response = client.get(f"/api/decode/jobs/{payload['id']}/download")
    assert download_response.status_code == 200
    assert download_response.headers["content-type"] == "application/zip"
    assert _zip_names(download_response.content) == [TREND_WORKBOOK_NAME]

    repeat_download = client.get(f"/api/decode/jobs/{payload['id']}/download")
    assert repeat_download.status_code == 410
    assert repeat_download.json()["detail"] == "Decode archive already downloaded. Re-upload files to download again."


def test_decode_job_download_returns_conflict_while_processing() -> None:
    job = decode_job_service.create_decode_job(selected_families=["Trend Chart"])

    response = client.get(f"/api/decode/jobs/{job.id}/download")

    assert response.status_code == 409
    assert response.json()["detail"] == "Decode job is not ready yet"


def test_decode_job_download_returns_conflict_when_failed() -> None:
    job = decode_job_service.create_decode_job(selected_families=["Trend Chart"])
    decode_job_service.update_decode_job(
        job.id,
        status="error",
        progress_percent=100,
        phase="Decode failed",
        detail="Unable to finish workbook export",
        error_message="simulated failure",
    )

    response = client.get(f"/api/decode/jobs/{job.id}/download")

    assert response.status_code == 409
    assert response.json()["detail"] == "simulated failure"


def test_decode_job_download_returns_gone_when_archive_expires(monkeypatch: pytest.MonkeyPatch) -> None:
    job = decode_job_service.create_decode_job(selected_families=["Trend Chart"])
    decode_job_service.update_decode_job(
        job.id,
        status="completed",
        progress_percent=100,
        phase="Ready to download",
        detail="Decoded workbooks are ready",
        archive_bytes=b"archive",
    )

    expired_now = job.updated_at + decode_job_service.JOB_RETENTION + timedelta(seconds=1)
    monkeypatch.setattr(decode_job_service, "utcnow", lambda: expired_now)

    response = client.get(f"/api/decode/jobs/{job.id}/download")

    assert response.status_code == 410
    assert response.json()["detail"] == "Decode archive expired from server memory. Re-upload files to download again."


def test_build_decode_export_archive_does_not_mutate_database_rows(data_dir) -> None:
    with _build_session() as db:
        _seed_decode_irrelevant_rows(db)
        counts_before = _table_counts(db)

        archive = build_decode_export_archive(
            trend_data=_file_bytes(data_dir, "Records/TrendChartRecord.data"),
            trend_index=_file_bytes(data_dir, "Records/TrendChartRecord.Index"),
            nibp_data=_file_bytes(data_dir, "Records/NibpRecord.data"),
            nibp_index=_file_bytes(data_dir, "Records/NibpRecord.Index"),
            timezone_name="Asia/Singapore",
        )

        counts_after = _table_counts(db)

    assert len(archive) > 0
    assert counts_before == counts_after


def test_decode_export_workbooks_keep_only_timestamp_utc_columns(data_dir) -> None:
    archive = build_decode_export_archive(
        trend_data=_file_bytes(data_dir, "Records/TrendChartRecord.data"),
        trend_index=_file_bytes(data_dir, "Records/TrendChartRecord.Index"),
        nibp_data=_file_bytes(data_dir, "Records/NibpRecord.data"),
        nibp_index=_file_bytes(data_dir, "Records/NibpRecord.Index"),
        timezone_name="Asia/Singapore",
    )

    trend_headers = _sheet_headers_from_zip(archive, TREND_WORKBOOK_NAME, "trend_frames")
    nibp_headers = _sheet_headers_from_zip(archive, NIBP_WORKBOOK_NAME, "Decoded_All")
    index_headers = _sheet_headers_from_zip(archive, TREND_WORKBOOK_NAME, "index_entries")

    assert "timestamp_utc" in trend_headers
    assert "timestamp_local" not in trend_headers
    assert "timestamp_utc" in nibp_headers
    assert "timestamp_local" not in nibp_headers
    assert "timestamp_utc" in index_headers
    assert "timestamp_local" not in index_headers


def test_decode_job_purge_enforces_job_count_cap() -> None:
    with decode_job_service._lock:
        now = decode_job_service.utcnow()
        for index in range(decode_job_service.MAX_RETAINED_JOBS + 2):
            job = decode_job_service.DecodeJob(
                id=f"job-{index}",
                status="completed",
                progress_percent=100,
                phase="Ready",
                detail=None,
                created_at=now - timedelta(seconds=index + 1),
                updated_at=now - timedelta(seconds=index + 1),
            )
            decode_job_service._jobs[job.id] = job

        decode_job_service._purge_expired_jobs()

        assert len(decode_job_service._jobs) == decode_job_service.MAX_RETAINED_JOBS
        assert "job-21" not in decode_job_service._jobs
        assert "job-20" not in decode_job_service._jobs


def test_decode_job_purge_drops_archives_when_byte_cap_is_exceeded(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(decode_job_service, "MAX_RETAINED_ARCHIVE_BYTES", 10)
    with decode_job_service._lock:
        now = decode_job_service.utcnow()
        for index in range(3):
            job = decode_job_service.DecodeJob(
                id=f"archive-job-{index}",
                status="completed",
                progress_percent=100,
                phase="Ready",
                detail=None,
                created_at=now - timedelta(seconds=index + 1),
                updated_at=now - timedelta(seconds=index + 1),
                archive_bytes=b"abcdef",
            )
            decode_job_service._jobs[job.id] = job

        decode_job_service._purge_expired_jobs()

        remaining_archive_bytes = sum(len(job.archive_bytes or b"") for job in decode_job_service._jobs.values())
        assert remaining_archive_bytes <= decode_job_service.MAX_RETAINED_ARCHIVE_BYTES
        assert decode_job_service._jobs["archive-job-2"].archive_bytes is None
        assert decode_job_service._jobs["archive-job-0"].archive_bytes == b"abcdef"
