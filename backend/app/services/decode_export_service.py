from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
import zipfile
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ..config import settings
from ..parsers.index_parser import ParsedIndex, parse_index_bytes
from ..parsers.nibp_parser import CHANNEL_COUNT as NIBP_CHANNEL_COUNT
from ..parsers.nibp_parser import nibp_channel_name, parse_nibp_frames
from ..parsers.trend_parser import CHANNEL_COUNT as TREND_CHANNEL_COUNT
from ..parsers.trend_parser import FRAME_SIZE as TREND_FRAME_SIZE
from ..parsers.trend_parser import PAYLOAD_SIZE as TREND_PAYLOAD_SIZE
from ..parsers.trend_parser import parse_trend_frames
from ..utils import coerce_utc

TREND_WORKBOOK_NAME = "TrendChartRecord_decoded.xlsx"
NIBP_WORKBOOK_NAME = "NibpRecord_decoded.xlsx"
ProgressCallback = Callable[[int, str, str | None], None]


def _timestamp_unix(timestamp: datetime) -> int:
    return int(coerce_utc(timestamp).timestamp())


def _timestamp_utc_text(timestamp: datetime) -> str:
    return coerce_utc(timestamp).strftime("%Y-%m-%d %H:%M:%S")


def _is_current_2025_2026(timestamp: datetime) -> bool:
    normalized = coerce_utc(timestamp)
    return normalized.year in {2025, 2026}


def _cell_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (int, float, bool, str)):
        return value
    if hasattr(value, "value") and isinstance(getattr(value, "value"), str):
        return getattr(value, "value")
    if isinstance(value, datetime):
        return _timestamp_utc_text(value)
    return str(value)


def _build_workbook(
    sheet_specs: list[tuple[str, list[str], list[list[object]]]],
    *,
    progress: Callable[[float], None] | None = None,
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    total_rows = sum(len(rows) for _sheet_name, _headers, rows in sheet_specs)
    written_rows = 0

    for sheet_name, headers, rows in sheet_specs:
        worksheet = workbook.create_sheet(title=sheet_name[:31])
        worksheet.append(headers)
        max_lengths = [len(str(header)) for header in headers]

        for row in rows:
            normalized_row = [_cell_value(value) for value in row]
            worksheet.append(normalized_row)
            for index, value in enumerate(normalized_row):
                length = len("" if value is None else str(value))
                if length > max_lengths[index]:
                    max_lengths[index] = length
            written_rows += 1
            if progress is not None and (written_rows % 250 == 0 or written_rows == total_rows):
                progress(written_rows / max(1, total_rows))

        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)

        for column_index, width in enumerate(max_lengths, start=1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 10), 48)

    buffer = BytesIO()
    workbook.save(buffer)
    if progress is not None:
        progress(1.0)
    return buffer.getvalue()


def _index_sheet_rows(parsed_index: ParsedIndex) -> list[list[object]]:
    rollback_set = set(parsed_index.rollback_indices)
    rows: list[list[object]] = []
    for entry in parsed_index.entries:
        rows.append(
            [
                entry.frame_index,
                entry.flags,
                entry.ptr,
                entry.ts_lo,
                entry.ts_hi,
                entry.unix_seconds,
                _timestamp_utc_text(entry.timestamp_utc),
                entry.frame_index in rollback_set,
            ]
        )
    return rows


def _trend_channel_headers() -> list[str]:
    return [f"ch{channel_index:02d}_be_u16_o{channel_index * 2:03d}" for channel_index in range(TREND_CHANNEL_COUNT)]


def _nibp_channel_headers() -> list[str]:
    return [nibp_channel_name(channel_index) for channel_index in range(NIBP_CHANNEL_COUNT)]


def _build_trend_workbook(
    *,
    trend_data: bytes,
    trend_index: bytes,
    timezone_name: str,
    progress_callback: Callable[[float, str], None] | None = None,
) -> bytes:
    if progress_callback is not None:
        progress_callback(0.05, "Reading Trend Chart index")
    parsed_index = parse_index_bytes(trend_index, trend_data, require_monotonic=False)
    if progress_callback is not None:
        progress_callback(0.2, "Parsing Trend Chart frames")
    frames = parse_trend_frames(trend_data, parsed_index, settings.invalid_u16_set)

    decode_notes_rows = [
        ["field", "value"],
        ["file_type", "TrendChartRecord reverse-engineering notes"],
        ["index_header_size", 16],
        ["index_entry_size", 16],
        ["index_trailer_size", 6],
        ["data_frame_size", TREND_FRAME_SIZE],
        ["data_payload_size", TREND_PAYLOAD_SIZE],
        ["channel_count", TREND_CHANNEL_COUNT],
        ["value_endianness", "big-endian u16"],
        ["invalid_u16_values", ",".join(str(value) for value in sorted(settings.invalid_u16_set))],
        ["decoded_frame_count", len(frames)],
        ["index_entry_count", len(parsed_index.entries)],
        ["index_timestamp_rollback_count", len(parsed_index.rollback_indices)],
        ["index_trailer_hex", parsed_index.trailer.hex()],
    ]

    index_header_rows = [
        ["field", "value"],
        ["version", parsed_index.header.version],
        ["buffer_size", parsed_index.header.buffer_size],
        ["payload_size", parsed_index.header.payload_size],
        ["padding", parsed_index.header.padding],
        ["entry_count", len(parsed_index.entries)],
        ["trailer_hex", parsed_index.trailer.hex()],
    ]

    trend_frame_headers = ["frame_index", "ptr", "timestamp_utc", *_trend_channel_headers(), "frame_tail_hex"]
    trend_frame_rows: list[list[object]] = []
    for frame, entry in zip(frames, parsed_index.entries, strict=True):
        frame_offset = frame.frame_index * TREND_FRAME_SIZE
        frame_tail_hex = trend_data[frame_offset + TREND_PAYLOAD_SIZE : frame_offset + TREND_FRAME_SIZE].hex()
        trend_frame_rows.append(
            [
                frame.frame_index,
                entry.ptr,
                _timestamp_utc_text(frame.timestamp),
                *frame.values,
                frame_tail_hex,
            ]
        )

    index_entry_headers = [
        "frame_index",
        "flags",
        "ptr",
        "ts_lo",
        "ts_hi",
        "unix_seconds",
        "timestamp_utc",
        "is_rollback",
    ]

    if progress_callback is not None:
        progress_callback(0.45, "Preparing Trend Chart workbook rows")
    return _build_workbook(
        [
            ("decode_notes", decode_notes_rows[0], decode_notes_rows[1:]),
            ("index_header", index_header_rows[0], index_header_rows[1:]),
            ("index_entries", index_entry_headers, _index_sheet_rows(parsed_index)),
            ("trend_frames", trend_frame_headers, trend_frame_rows),
        ],
        progress=(
            (lambda fraction: progress_callback(0.45 + (fraction * 0.55), "Writing Trend Chart workbook"))
            if progress_callback is not None
            else None
        ),
    )


def _build_nibp_workbook(
    *,
    nibp_data: bytes,
    nibp_index: bytes,
    timezone_name: str,
    progress_callback: Callable[[float, str], None] | None = None,
) -> bytes:
    if progress_callback is not None:
        progress_callback(0.05, "Reading NIBP index")
    parsed_index = parse_index_bytes(nibp_index, nibp_data, require_monotonic=False)
    if progress_callback is not None:
        progress_callback(0.2, "Parsing NIBP records")
    frames = parse_nibp_frames(nibp_data, parsed_index, settings.invalid_u16_set)

    all_headers = [
        "record_index",
        "index_flags",
        "index_ptr",
        "timestamp_unix",
        "timestamp_utc",
        "has_measurement",
        "bp_systolic_inferred",
        "bp_mean_inferred",
        "bp_diastolic_inferred",
        *_nibp_channel_headers(),
    ]

    all_rows: list[list[object]] = []
    raw_hex_rows: list[list[object]] = []

    for frame, entry in zip(frames, parsed_index.entries, strict=True):
        row = [
            frame.frame_index,
            entry.flags,
            entry.ptr,
            _timestamp_unix(frame.timestamp),
            _timestamp_utc_text(frame.timestamp),
            frame.has_measurement,
            frame.channel_values.get("bp_systolic_inferred"),
            frame.channel_values.get("bp_mean_inferred"),
            frame.channel_values.get("bp_diastolic_inferred"),
            *[frame.channel_values.get(nibp_channel_name(channel_index)) for channel_index in range(NIBP_CHANNEL_COUNT)],
        ]
        all_rows.append(row)

        frame_offset = frame.frame_index * TREND_FRAME_SIZE
        raw_hex_rows.append(
            [
                frame.frame_index,
                _timestamp_utc_text(frame.timestamp),
                nibp_data[frame_offset : frame_offset + TREND_FRAME_SIZE].hex().upper(),
            ]
        )

    current_rows = [row for row, frame in zip(all_rows, frames, strict=True) if _is_current_2025_2026(frame.timestamp)]
    bp_rows = [
        row
        for row, frame in zip(all_rows, frames, strict=True)
        if all(
            frame.channel_values.get(key) is not None
            for key in ("bp_systolic_inferred", "bp_mean_inferred", "bp_diastolic_inferred")
        )
    ]

    note_headers = ["section", "field", "value", "notes"]
    note_rows = [
        ["Inputs", "data_record_size", TREND_FRAME_SIZE, "Each NIBP record is fixed-width at 124 bytes."],
        ["Inputs", "data_record_count", len(frames), "Derived from data file length / 124."],
        ["Inputs", "index_header_size", 16, "Index file uses the standard 16-byte header in the backend parser."],
        ["Inputs", "index_entry_size", 16, "Each index entry is 16 bytes."],
        ["Decode", "bp_systolic_inferred", "inferred", "Mapped from the backend NIBP parser inferred systolic field."],
        ["Decode", "bp_mean_inferred", "inferred", "Mapped from the backend NIBP parser inferred mean field."],
        ["Decode", "bp_diastolic_inferred", "inferred", "Mapped from the backend NIBP parser inferred diastolic field."],
        ["Quality", "current_2025_2026_records", len(current_rows), "Rows whose UTC year falls in 2025 or 2026."],
        ["Quality", "records_with_bp_triplet", len(bp_rows), "Rows where all three inferred BP values are present."],
    ]

    raw_hex_headers = ["record_index", "timestamp_utc", "raw_hex"]
    index_headers = [
        "frame_index",
        "flags",
        "ptr",
        "ts_lo",
        "ts_hi",
        "unix_seconds",
        "timestamp_utc",
        "is_rollback",
    ]

    if progress_callback is not None:
        progress_callback(0.45, "Preparing NIBP workbook rows")
    return _build_workbook(
        [
            ("Decoded_All", all_headers, all_rows),
            ("Decoded_Current", all_headers, current_rows),
            ("Decoded_With_BP", all_headers, bp_rows),
            ("Index", index_headers, _index_sheet_rows(parsed_index)),
            ("Notes", note_headers, note_rows),
            ("RawHex", raw_hex_headers, raw_hex_rows),
        ],
        progress=(
            (lambda fraction: progress_callback(0.45 + (fraction * 0.55), "Writing NIBP workbook"))
            if progress_callback is not None
            else None
        ),
    )


def build_decode_export_archive(
    *,
    trend_data: bytes | None = None,
    trend_index: bytes | None = None,
    nibp_data: bytes | None = None,
    nibp_index: bytes | None = None,
    timezone_name: str = "UTC",
    progress_callback: ProgressCallback | None = None,
) -> bytes:
    selected_families = [
        ("Trend Chart", TREND_WORKBOOK_NAME, trend_data is not None and trend_index is not None),
        ("NIBP", NIBP_WORKBOOK_NAME, nibp_data is not None and nibp_index is not None),
    ]
    active_families = [(label, filename) for label, filename, is_selected in selected_families if is_selected]

    def report(progress: int, phase: str, detail: str | None = None) -> None:
        if progress_callback is not None:
            progress_callback(progress, phase, detail)

    archive = BytesIO()
    with zipfile.ZipFile(archive, mode="w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        completed_families = 0

        if trend_data is not None and trend_index is not None:
            family_count = max(1, len(active_families))
            zip_file.writestr(
                TREND_WORKBOOK_NAME,
                _build_trend_workbook(
                    trend_data=trend_data,
                    trend_index=trend_index,
                    timezone_name=timezone_name,
                    progress_callback=(
                        lambda fraction, phase: report(
                            round(((completed_families + fraction) / family_count) * 95),
                            phase,
                            "Decoding Trend Chart files",
                        )
                    ),
                ),
            )
            completed_families += 1
        if nibp_data is not None and nibp_index is not None:
            family_count = max(1, len(active_families))
            zip_file.writestr(
                NIBP_WORKBOOK_NAME,
                _build_nibp_workbook(
                    nibp_data=nibp_data,
                    nibp_index=nibp_index,
                    timezone_name=timezone_name,
                    progress_callback=(
                        lambda fraction, phase: report(
                            round(((completed_families + fraction) / family_count) * 95),
                            phase,
                            "Decoding NIBP files",
                        )
                    ),
                ),
            )
            completed_families += 1

    report(98, "Packaging archive", "Preparing the ZIP download")
    return archive.getvalue()
