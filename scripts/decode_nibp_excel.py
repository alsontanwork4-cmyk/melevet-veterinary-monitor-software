from __future__ import annotations

import argparse
import math
import struct
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


RECORD_SIZE = 124
INDEX_HEADER_SIZE = 22
INDEX_ENTRY_SIZE = 16
CURRENT_DATA_START_UNIX = int(datetime(2025, 1, 1, tzinfo=timezone.utc).timestamp())
CURRENT_DATA_END_UNIX = int(datetime(2027, 1, 1, tzinfo=timezone.utc).timestamp())
LOCAL_TZ = ZoneInfo("Asia/Singapore")


@dataclass
class IndexEntry:
    record_index: int
    slot: int
    timestamp_unix: int
    sequence: int
    index_code: int


def u16le(blob: bytes, offset: int) -> int:
    return struct.unpack_from("<H", blob, offset)[0]


def u32le(blob: bytes, offset: int) -> int:
    return struct.unpack_from("<I", blob, offset)[0]


def u32be(blob: bytes, offset: int) -> int:
    return struct.unpack_from(">I", blob, offset)[0]


def f32be_or_none(blob: bytes, offset: int) -> float | None:
    value = struct.unpack_from(">f", blob, offset)[0]
    if math.isnan(value) or math.isinf(value):
        return None
    return round(value, 2)


def u32be_or_none(blob: bytes, offset: int) -> int | None:
    value = u32be(blob, offset)
    if value == 0xFFFFFFFF:
        return None
    return value


def iso_utc(ts: int) -> str:
    return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def iso_local(ts: int) -> str:
    return datetime.fromtimestamp(ts, tz=timezone.utc).astimezone(LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S")


def record_hex(record: bytes) -> str:
    return record.hex().upper()


def parse_index(index_bytes: bytes) -> tuple[dict[str, int], list[IndexEntry]]:
    if len(index_bytes) < INDEX_HEADER_SIZE:
        raise ValueError("Index file is too small to contain the expected header")
    if (len(index_bytes) - INDEX_HEADER_SIZE) % INDEX_ENTRY_SIZE != 0:
        raise ValueError(
            f"Unexpected index file size {len(index_bytes)}; trailing bytes do not fit 16-byte entries"
        )

    header = {
        "header_u32le_0": u32le(index_bytes, 0),
        "header_u32le_4": u32le(index_bytes, 4),
        "header_u32le_8": u32le(index_bytes, 8),
        "header_u32le_12": u32le(index_bytes, 12),
        "header_u16le_16": u16le(index_bytes, 16),
        "header_u16le_18": u16le(index_bytes, 18),
        "header_u16le_20": u16le(index_bytes, 20),
    }

    entries: list[IndexEntry] = []
    for record_index in range((len(index_bytes) - INDEX_HEADER_SIZE) // INDEX_ENTRY_SIZE):
        offset = INDEX_HEADER_SIZE + record_index * INDEX_ENTRY_SIZE
        entries.append(
            IndexEntry(
                record_index=record_index,
                slot=u32le(index_bytes, offset),
                timestamp_unix=u32le(index_bytes, offset + 4),
                sequence=u32le(index_bytes, offset + 8),
                index_code=u32be(index_bytes, offset + 12),
            )
        )

    return header, entries


def build_rows(data_bytes: bytes, index_entries: list[IndexEntry]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if len(data_bytes) % RECORD_SIZE != 0:
        raise ValueError(f"Data file size {len(data_bytes)} is not a multiple of {RECORD_SIZE}")

    record_count = len(data_bytes) // RECORD_SIZE
    if record_count != len(index_entries):
        raise ValueError(
            f"Index/data mismatch: data has {record_count} records but index has {len(index_entries)} entries"
        )

    all_rows: list[dict[str, Any]] = []
    raw_rows: list[dict[str, Any]] = []

    for record_index in range(record_count):
        offset = record_index * RECORD_SIZE
        record = data_bytes[offset : offset + RECORD_SIZE]
        index_entry = index_entries[record_index]

        timestamp_unix = u32le(record, 2)
        sequence_le = u32le(record, 6)
        timestamp_duplicate_be = u32be(record, 10)
        sequence_duplicate_be = u32be(record, 14)

        metric_a = u32be_or_none(record, 26)
        metric_b = u32be_or_none(record, 30)
        bp_systolic = f32be_or_none(record, 39)
        bp_diastolic = f32be_or_none(record, 43)
        bp_mean = f32be_or_none(record, 47)
        aux_float = f32be_or_none(record, 69)

        is_current = CURRENT_DATA_START_UNIX <= timestamp_unix < CURRENT_DATA_END_UNIX
        has_bp_triplet = all(value is not None for value in (bp_systolic, bp_diastolic, bp_mean))

        all_rows.append(
            {
                "record_index": record_index,
                "index_slot": index_entry.slot,
                "slot_matches_position": index_entry.slot == record_index + 1,
                "byte_offset": offset,
                "timestamp_unix": timestamp_unix,
                "timestamp_utc": iso_utc(timestamp_unix),
                "timestamp_singapore": iso_local(timestamp_unix),
                "index_timestamp_unix": index_entry.timestamp_unix,
                "index_timestamp_utc": iso_utc(index_entry.timestamp_unix),
                "index_timestamp_matches_data": index_entry.timestamp_unix == timestamp_unix,
                "sequence": sequence_le,
                "sequence_duplicate_be": sequence_duplicate_be,
                "sequence_matches_duplicate": sequence_le == sequence_duplicate_be,
                "timestamp_duplicate_be": timestamp_duplicate_be,
                "timestamp_duplicate_matches_data": timestamp_duplicate_be == timestamp_unix,
                "record_flags_u16le": u16le(record, 0),
                "status_flag_u8": record[34],
                "metric_a_u32be": metric_a,
                "metric_b_u32be": metric_b,
                "bp_systolic_inferred": bp_systolic,
                "bp_diastolic_inferred": bp_diastolic,
                "bp_mean_inferred": bp_mean,
                "aux_float_69_inferred": aux_float,
                "constant_float_52": f32be_or_none(record, 52),
                "constant_float_79": f32be_or_none(record, 79),
                "index_code": index_entry.index_code,
                "has_bp_triplet": has_bp_triplet,
                "is_current_2025_2026": is_current,
            }
        )

        raw_rows.append(
            {
                "record_index": record_index,
                "timestamp_utc": iso_utc(timestamp_unix),
                "raw_hex": record_hex(record),
            }
        )

    return all_rows, raw_rows


def notes_rows(index_header: dict[str, int], all_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    current_count = sum(1 for row in all_rows if row["is_current_2025_2026"])
    bp_count = sum(1 for row in all_rows if row["has_bp_triplet"])
    stale_count = len(all_rows) - current_count

    return [
        {"section": "Inputs", "field": "data_record_size", "value": RECORD_SIZE, "notes": "Each data record is fixed-width at 124 bytes."},
        {"section": "Inputs", "field": "data_record_count", "value": len(all_rows), "notes": "Derived from data file length / 124."},
        {"section": "Inputs", "field": "index_header_size", "value": INDEX_HEADER_SIZE, "notes": "Index file has a 22-byte header before 16-byte entries."},
        {"section": "Inputs", "field": "index_entry_size", "value": INDEX_ENTRY_SIZE, "notes": "Each index entry is 16 bytes."},
        {"section": "Header", "field": "header_u32le_0", "value": index_header["header_u32le_0"], "notes": "Unknown header value."},
        {"section": "Header", "field": "header_u32le_4", "value": index_header["header_u32le_4"], "notes": "Unknown header value; observed as 5000 in this file."},
        {"section": "Header", "field": "header_u32le_8", "value": index_header["header_u32le_8"], "notes": "Unknown header value."},
        {"section": "Decode", "field": "timestamp_unix", "value": "decoded", "notes": "Stored at data offset +2 as little-endian Unix seconds."},
        {"section": "Decode", "field": "sequence", "value": "decoded", "notes": "Stored at data offset +6 as little-endian uint32 and duplicated at +14 as big-endian uint32."},
        {"section": "Decode", "field": "bp_systolic_inferred", "value": "inferred", "notes": "Big-endian float at offset +39. Range and grouping strongly match systolic pressure."},
        {"section": "Decode", "field": "bp_diastolic_inferred", "value": "inferred", "notes": "Big-endian float at offset +43. Range and grouping strongly match diastolic pressure."},
        {"section": "Decode", "field": "bp_mean_inferred", "value": "inferred", "notes": "Big-endian float at offset +47. Range and grouping strongly match mean arterial pressure."},
        {"section": "Decode", "field": "metric_a_u32be", "value": "partially decoded", "notes": "Big-endian uint32 at offset +26. Observed range 50-100. Likely a saturation/quality metric, but not proven from the binary alone."},
        {"section": "Decode", "field": "metric_b_u32be", "value": "partially decoded", "notes": "Big-endian uint32 at offset +30. Observed range 34-228. Likely a pulse-rate-type field, but not proven from the binary alone."},
        {"section": "Decode", "field": "aux_float_69_inferred", "value": "partially decoded", "notes": "Big-endian float at offset +69. Observed range 21.2-82.22. Meaning is still unknown."},
        {"section": "Quality", "field": "current_2025_2026_records", "value": current_count, "notes": "Records whose decoded timestamp falls in 2025-2026."},
        {"section": "Quality", "field": "stale_or_buffer_records", "value": stale_count, "notes": "A small number of records decode to around year 2000 and appear to be circular-buffer leftovers."},
        {"section": "Quality", "field": "records_with_bp_triplet", "value": bp_count, "notes": "Rows where all three inferred BP floats are present and finite."},
    ]


def autosize_and_format(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.auto_filter.ref = ws.dimensions

        for column_cells in ws.columns:
            max_len = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 48)
    wb.save(workbook_path)


def write_workbook(
    data_path: Path,
    index_path: Path,
    output_path: Path,
) -> None:
    data_bytes = data_path.read_bytes()
    index_bytes = index_path.read_bytes()

    index_header, index_entries = parse_index(index_bytes)
    all_rows, raw_rows = build_rows(data_bytes, index_entries)

    index_rows = [
        {
            "record_index": entry.record_index,
            "index_slot": entry.slot,
            "timestamp_unix": entry.timestamp_unix,
            "timestamp_utc": iso_utc(entry.timestamp_unix),
            "timestamp_singapore": iso_local(entry.timestamp_unix),
            "sequence": entry.sequence,
            "index_code": entry.index_code,
        }
        for entry in index_entries
    ]

    current_rows = [row for row in all_rows if row["is_current_2025_2026"]]
    bp_rows = [row for row in all_rows if row["has_bp_triplet"]]

    notes = notes_rows(index_header, all_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(all_rows).to_excel(writer, sheet_name="Decoded_All", index=False)
        pd.DataFrame(current_rows).to_excel(writer, sheet_name="Decoded_Current", index=False)
        pd.DataFrame(bp_rows).to_excel(writer, sheet_name="Decoded_With_BP", index=False)
        pd.DataFrame(index_rows).to_excel(writer, sheet_name="Index", index=False)
        pd.DataFrame(notes).to_excel(writer, sheet_name="Notes", index=False)
        pd.DataFrame(raw_rows).to_excel(writer, sheet_name="RawHex", index=False)

    autosize_and_format(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Decode linked NIBP data/index files into an Excel workbook.")
    parser.add_argument(
        "--data",
        default=r"C:\Users\Alson\Desktop\NibpRecord.data",
        help="Path to NibpRecord.data",
    )
    parser.add_argument(
        "--index",
        default=r"C:\Users\Alson\Desktop\NibpRecord.Index",
        help="Path to NibpRecord.Index",
    )
    parser.add_argument(
        "--output",
        default=r"C:\Users\Alson\Desktop\NibpRecord_decoded.xlsx",
        help="Path to output .xlsx workbook",
    )
    args = parser.parse_args()

    data_path = Path(args.data)
    index_path = Path(args.index)
    output_path = Path(args.output)

    if not data_path.exists():
        raise FileNotFoundError(f"Data file not found: {data_path}")
    if not index_path.exists():
        raise FileNotFoundError(f"Index file not found: {index_path}")

    write_workbook(data_path=data_path, index_path=index_path, output_path=output_path)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
